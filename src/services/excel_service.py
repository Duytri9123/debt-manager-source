# -*- coding: utf-8 -*-
"""
Excel Service - Import/Export functionality
Mirrors: Laravel Excel import/export in B2BOrderController
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional
import os


class ExcelService:
    """Service for Excel operations"""
    
    def __init__(self):
        # Colors matching Laravel export
        self.FONT_NAME = 'Times New Roman'
        self.YELLOW = 'FFD700'
        self.LYELLOW = 'FFFACD'
        self.RED = 'FF0000'
        self.BORDER = '000000'
        
    def export_order_to_excel(self, order: Dict, items: List[Dict], output_path: str) -> str:
        """Export single order to Excel (mirrors Laravel exportExcel)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Don hang'
        
        # Styles
        fill_yellow = PatternFill(start_color=self.YELLOW, end_color=self.YELLOW, fill_type='solid')
        fill_lyellow = PatternFill(start_color=self.LYELLOW, end_color=self.LYELLOW, fill_type='solid')
        border_thin = Border(
            left=Side(style='thin', color=self.BORDER),
            right=Side(style='thin', color=self.BORDER),
            top=Side(style='thin', color=self.BORDER),
            bottom=Side(style='thin', color=self.BORDER)
        )
        align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='top', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Title
        ws.merge_cells('A1:J1')
        title = f"ĐƠN HÀNG: {order.get('order_name', order.get('order_number', '')).upper()}"
        ws['A1'] = title
        ws['A1'].font = Font(name=self.FONT_NAME, bold=True, size=16, color='1F3864')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        
        # Order info
        order_date = order.get('created_at', '')[:10] if order.get('created_at') else ''
        delivery_date = order.get('delivery_date', '—')
        
        info_data = [
            (2, 'Mã đơn:', order.get('order_number', ''), 'Ngày đặt:', order_date),
            (3, 'Khách hàng:', order.get('customer_name', ''), 'Ngày xuất:', delivery_date),
            (4, 'SĐT:', order.get('customer_phone', ''), 'Email:', order.get('customer_email', '')),
        ]
        
        for row, la, va, lb, vb in info_data:
            ws[f'A{row}'] = la
            ws[f'A{row}'].font = Font(name=self.FONT_NAME, bold=True, size=10)
            ws[f'B{row}'] = va
            ws[f'B{row}'].font = Font(name=self.FONT_NAME, size=10)
            ws[f'F{row}'] = lb
            ws[f'F{row}'].font = Font(name=self.FONT_NAME, bold=True, size=10)
            ws[f'G{row}'] = vb
            ws[f'G{row}'].font = Font(name=self.FONT_NAME, size=10)
            
        if order.get('shipping_address'):
            ws['A5'] = 'Địa chỉ:'
            ws['A5'].font = Font(name=self.FONT_NAME, bold=True, size=10)
            ws.merge_cells('B5:J5')
            ws['B5'] = order['shipping_address']
            ws['B5'].font = Font(name=self.FONT_NAME, size=10)
            
        note_row = 6 if order.get('shipping_address') else 5
        if order.get('notes'):
            ws[f'A{note_row}'] = 'Ghi chú:'
            ws[f'A{note_row}'].font = Font(name=self.FONT_NAME, bold=True, size=10)
            ws.merge_cells(f'B{note_row}:J{note_row}')
            ws[f'B{note_row}'] = order['notes']
            ws[f'B{note_row}'].font = Font(name=self.FONT_NAME, size=10)
            
        # Headers
        header_row = note_row + 2
        headers = [
            ('A', 'Năm', 6, align_center),
            ('B', 'TT', 6, align_center),
            ('C', 'MÔ TẢ CHI TIẾT', 42, align_center),
            ('D', 'MÃ HÀNG', 14, align_center),
            ('E', 'XUẤT XỨ', 10, align_center),
            ('F', 'ĐƠN VỊ', 10, align_center),
            ('G', 'SỐ LƯỢNG', 10, align_center),
            ('H', 'ĐƠN GIÁ', 16, align_right),
            ('I', 'THÀNH TIỀN', 16, align_right),
            ('J', 'GHI CHÚ', 18, align_left),
        ]
        
        for col, label, width, align in headers:
            ws.column_dimensions[col].width = width
            cell = ws[f'{col}{header_row}']
            cell.value = label
            cell.font = Font(name=self.FONT_NAME, bold=True, size=10)
            cell.fill = fill_yellow
            cell.alignment = align
            cell.border = border_thin
            
        ws.row_dimensions[header_row].height = 28
        
        # Data rows
        row = header_row + 1
        year = order.get('created_at', '')[:4] if order.get('created_at') else datetime.now().year
        
        for item in items:
            attrs = item.get('variant_attributes', '{}')
            try:
                import json
                attrs_dict = json.loads(attrs) if isinstance(attrs, str) else attrs
            except:
                attrs_dict = {}
                
            is_category = attrs_dict.get('type') == 'category'
            note = attrs_dict.get('note', '')
            
            if is_category:
                ws[f'A{row}'] = year
                ws.merge_cells(f'C{row}:I{row}')
                ws[f'C{row}'] = item['product_name'].upper()
                
                for col in range(ord('A'), ord('J') + 1):
                    cell = ws[f'{chr(col)}{row}']
                    cell.font = Font(name=self.FONT_NAME, bold=True, size=10)
                    cell.fill = fill_lyellow
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = border_thin
                    
                ws[f'C{row}'].alignment = align_left
            else:
                line_total = item.get('line_total', 0) or (item.get('price', 0) * item.get('quantity', 0))
                
                ws[f'A{row}'] = year
                ws[f'B{row}'] = '+'
                ws[f'C{row}'] = item.get('product_name', '')
                ws[f'D{row}'] = item.get('variant_sku', '')
                ws[f'E{row}'] = item.get('origin', '')
                ws[f'F{row}'] = item.get('unit', '')
                ws[f'G{row}'] = int(item.get('quantity', 0))
                ws[f'H{row}'] = item.get('price', 0)
                ws[f'I{row}'] = line_total
                ws[f'J{row}'] = note
                
                for col in range(ord('A'), ord('J') + 1):
                    cell = ws[f'{chr(col)}{row}']
                    cell.font = Font(name=self.FONT_NAME, size=10)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = border_thin
                    
                ws[f'A{row}'].alignment = align_center
                ws[f'B{row}'].alignment = align_center
                ws[f'G{row}'].alignment = align_center
                ws[f'H{row}'].alignment = align_right
                ws[f'I{row}'].alignment = align_right
                ws[f'H{row}'].number_format = '#,##0'
                ws[f'I{row}'].number_format = '#,##0'
                
                if note:
                    ws[f'J{row}'].font = Font(name=self.FONT_NAME, size=10, color=self.RED)
                    
            ws.row_dimensions[row].height = -1
            row += 1
            
        # Summary rows
        row += 1
        subtotal = order.get('subtotal', 0)
        grand_total = order.get('grand_total', 0)
        tax = grand_total - subtotal
        tax_rate = round(tax / subtotal * 100) if subtotal > 0 else 0
        
        summary_data = [
            ('TỔNG GIÁ TRỊ TRƯỚC THUẾ', subtotal, 'F3F4F6'),
            (f'THUẾ GTGT {tax_rate}%', tax, 'F3F4F6'),
            ('TỔNG GIÁ TRỊ SAU THUẾ', grand_total, self.YELLOW),
        ]
        
        for label, value, bg_color in summary_data:
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = label
            ws[f'I{row}'] = value
            
            for col in range(ord('A'), ord('J') + 1):
                cell = ws[f'{chr(col)}{row}']
                cell.font = Font(name=self.FONT_NAME, bold=True, size=10)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='right', vertical='top')
                cell.border = border_thin
                
            ws[f'I{row}'].number_format = '#,##0'
            row += 1
            
        # Save
        wb.save(output_path)
        return output_path
        
    def import_orders_from_excel(self, file_path: str) -> Dict:
        """Import B2B orders from Excel (mirrors Laravel importProcess)"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        skip_sheets = ['foxz', 'TỔNG', 'HH+DS']
        imported = 0
        skipped = 0
        errors = []
        all_orders = []
        
        for sheet_name in wb.sheetnames:
            # Skip certain sheets
            if any(skip.lower() in sheet_name.lower() for skip in skip_sheets):
                continue
                
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            # Parse orders from this sheet
            orders = self._parse_orders(rows)
            all_orders.extend(orders)
            
        return {
            'orders': all_orders,
            'total': len(all_orders),
            'errors': errors
        }
        
    def _parse_orders(self, rows: List) -> List[Dict]:
        """Parse orders from Excel rows (mirrors Laravel parseOrders)"""
        orders = []
        current = None
        
        summary_markers = [
            'TỔNG GIÁ TRỊ TRƯỚC THUẾ',
            'THUẾ GTGT 10%',
            'TỔNG GIÁ TRỊ SAU THUẾ'
        ]
        
        for row in rows:
            # Pad row to ensure we have enough columns
            while len(row) < 17:
                row += (None,)
                
            c0 = self._to_string(row[0])
            c1 = row[1]
            c2 = self._to_string(row[2])
            
            # Skip empty rows
            if not any(v is not None and v != '' for v in row):
                continue
                
            # Check for header row (date pattern)
            is_header = self._is_date_header(c2)
            is_column_header = c2 in ['TT', 'STT']
            
            if is_header or is_column_header:
                customer = c0 or (current['customer'] if current else None)
                order_date = self._parse_date(c1) or (current['date'] if current else None)
                
                if customer and order_date:
                    key = f"{customer}|{order_date}"
                    if not current or current.get('_key') != key:
                        if current:
                            orders.append(current)
                        current = {
                            '_key': key,
                            'customer': customer,
                            'date': order_date,
                            'items': [],
                            'total_before': 0,
                            'tax': 0,
                            'grand_total': 0
                        }
                continue
                
            if not current:
                customer = c0
                order_date = self._parse_date(c1)
                if customer and order_date:
                    current = {
                        '_key': f"{customer}|{order_date}",
                        'customer': customer,
                        'date': order_date,
                        'items': [],
                        'total_before': 0,
                        'tax': 0,
                        'grand_total': 0
                    }
                else:
                    continue
                    
            # Check for customer change
            if c0 and c0 != current['customer']:
                new_date = self._parse_date(c1) or current['date']
                new_key = f"{c0}|{new_date}"
                if new_key != current['_key']:
                    orders.append(current)
                    current = {
                        '_key': new_key,
                        'customer': c0,
                        'date': new_date,
                        'items': [],
                        'total_before': 0,
                        'tax': 0,
                        'grand_total': 0
                    }
                    
            # Check for summary rows
            if c2 in summary_markers:
                amount = self._parse_money(row[9])
                if c2 == 'TỔNG GIÁ TRỊ TRƯỚC THUẾ':
                    current['total_before'] = amount
                elif c2 == 'THUẾ GTGT 10%':
                    current['tax'] = amount
                elif c2 == 'TỔNG GIÁ TRỊ SAU THUẾ':
                    current['grand_total'] = amount
                continue
                
            # Parse item row
            stt = self._to_string(row[2])
            if not stt.isdigit() and stt != '+':
                continue
                
            description = self._to_string(row[3])
            if not description:
                continue
                
            current['items'].append({
                'description': description,
                'product_code': self._to_string(row[4]),
                'origin': self._to_string(row[5]),
                'unit': self._to_string(row[6]),
                'quantity': self._parse_number(row[7]) or 1,
                'unit_price': self._parse_money(row[8]),
                'line_total': self._parse_money(row[9]),
                'note': self._to_string(row[10]),
                'cost_price': self._parse_money(row[11]),
                'selling_price': self._parse_money(row[12]),
                'business_pct': self._parse_money(row[13]),
                'profit_per_kg': self._parse_money(row[14]),
                'weight_kg': self._parse_number(row[15]),
                'total_profit': self._parse_money(row[16]),
            })
            
        if current:
            orders.append(current)
            
        # Filter orders with items
        return [o for o in orders if o.get('items')]
        
    def _to_string(self, value) -> str:
        """Convert value to string"""
        if value is None:
            return ''
        return str(value).strip()
        
    def _parse_money(self, value) -> float:
        """Parse money value"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(',', '').replace(' ', ''))
        except:
            return 0.0
            
    def _parse_number(self, value) -> float:
        """Parse number value"""
        return self._parse_money(value)
        
    def _parse_date(self, value) -> Optional[str]:
        """Parse date value (handles Excel dates and strings)"""
        if value is None:
            return None
            
        # Handle Excel datetime objects
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
            
        # Handle Excel serial numbers
        if isinstance(value, (int, float)) and value > 1000:
            try:
                # Excel dates are days since 1899-12-30
                excel_epoch = datetime(1899, 12, 30)
                date_obj = excel_epoch + timedelta(days=value)
                return date_obj.strftime('%Y-%m-%d')
            except:
                pass
                
        # Parse string dates
        s = str(value).strip()
        for fmt in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d']:
            try:
                from datetime import datetime as dt
                date_obj = dt.strptime(s, fmt)
                return date_obj.strftime('%Y-%m-%d')
            except:
                continue
                
        return None
        
    def _is_date_header(self, value: str) -> bool:
        """Check if value is a date header"""
        import re
        return bool(re.match(r'^\d{1,2}/\d{1,2}/\d{4}', value))
